"""
predict_pipeline_with_YAMNet.py - predict_pipeline.py + YAMNet crowd-noise scoring.

Adds a 'Crowd Noise Score' column to the Excel output.
Score = max YAMNet probability across crowd-related AudioSet classes
(Cheering, Crowd, Applause, Yelling, Shouting) in the audio window
covering each throw's TO segment + 3 seconds after.

In goalball, play is silent (players are blind). When a goal is scored
the game stops and the crowd/bench react freely — a high crowd-noise
score is therefore a goal-specific signal that the LSTM cannot see.

Requirements (in addition to predict_pipeline deps):
    pip install tensorflow tensorflow-hub

Falls back gracefully (score = 0.0) if TensorFlow is not installed
or if ffmpeg is not on PATH.
"""

import os, pickle, warnings, time, subprocess, wave, tempfile, threading
from datetime import datetime
import numpy as np, pandas as pd, cv2
from pathlib import Path
from ultralytics import YOLO
from sklearn.preprocessing import StandardScaler
import torch, torch.nn as nn
from torch.nn.utils.rnn import pack_padded_sequence, pad_packed_sequence

warnings.filterwarnings("ignore")

_T0 = time.time()
def log(msg):
    """Print msg prefixed with wall-clock HH:MM:SS and elapsed since start."""
    elapsed = time.time() - _T0
    em, es = divmod(int(elapsed), 60)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] (+{em:02d}:{es:02d}) {msg}")

# Non-blocking console hotkey detection. Windows-only via msvcrt (stdlib);
# graceful no-op on Linux/macOS so the script still runs there.
try:
    import sounddevice as sd
    _SOUNDDEVICE_AVAILABLE = True
except ImportError:
    _SOUNDDEVICE_AVAILABLE = False

try:
    import msvcrt
    def _console_h_pressed():
        if msvcrt.kbhit():
            try:    return msvcrt.getch().decode('utf-8', errors='ignore').lower() == 'h'
            except: return False
        return False
except ImportError:
    def _console_h_pressed():
        return False  # non-Windows: focus the preview window and press H there

# ===================================================================
#  CONFIG
# ===================================================================
# Paths resolve relative to this script so the repo works on any machine
# once it's cloned from GitHub. Override with env vars if your data lives
# elsewhere (e.g. set GOALBALL_TRAIN_ROOT=/path/to/Paralkympics2024).
SCRIPT_DIR   = Path(__file__).resolve().parent
WEIGHTS_DIR  = SCRIPT_DIR / "Model Weights"
YOLO_WEIGHTS = WEIGHTS_DIR / "best.pt"
LSTM_WEIGHTS = WEIGHTS_DIR / "final_model.pt"
SCALER_CACHE = WEIGHTS_DIR / "scaler.pkl"

# If True, after the Excel is written, replay each throw (from+to frames)
# with YOLO bounding boxes overlaid. Press any key to advance to the next
# throw; press 'q' to quit the viewer.
SHOW_THROWS = True

# If True, drop throws the LSTM predicted as a false-positive class (o0/g0/b0)
# before writing the Excel. Default is False because the Excel already exposes
# the raw 'Predicted Class' column (o1/o0/b1/g1/...) so you can filter manually.
# Flip to True if you want the script to remove o0/g0/b0 rows automatically.
DROP_FALSE_THROWS = False

# If True, merge "phantom split" duplicates: when YOLO misclassifies a mid-throw
# frame as the other team for a few frames, it creates two throws covering the
# same physical event. Detection (both must hold):
#   (1) B's FROM starts within MERGE_PHANTOM_GAP_SECONDS of A's FROM start
#       (YOLO re-opened a new FROM mid-flight, so both FROMs start at the same moment)
#   (2) same throwing side (i.e. same throwing team)
# Using FROM-start gap (not TO-end gap) avoids false positives when the ball stays
# visible across a missed intermediate throw and A's TO tracking extends far forward.
# Resolution: keep the LATER throw (its TO captures the actual outcome moment;
# the earlier throw's TO was truncated by the phantom FROM split).
MERGE_DUPLICATE_THROWS    = True
MERGE_PHANTOM_GAP_SECONDS = 3   # max seconds between A's FROM start and B's FROM start

# ===================================================================
#  YAMNET CONFIG
# ===================================================================
# Audio window around each throw's TO segment for crowd-noise scoring.
# YAMNET_POST_SECONDS: extra seconds captured AFTER the TO segment ends
# to catch the crowd reaction that follows a goal.
YAMNET_POST_SECONDS = 3.0
# AudioSet class names to include in the crowd-noise score (case-insensitive
# substring match against YAMNet's 521-class label set).
# Stems matched case-insensitively as substrings against YAMNet's display_name column.
# Verified against: github.com/tensorflow/models/.../yamnet_class_map.csv
# Matched classes: Shout(6), Yell(9), Screaming(11), Cheering(61), Applause(62),
#                  Crowd(64), Whoop(8)
YAMNET_CROWD_KEYWORDS = ["cheer", "applause", "crowd", "yell", "shout", "whoop", "scream"]

# Used ONLY on first run to (re-)fit the StandardScaler, then cached.
# After scaler.pkl exists these paths are not read again, so users who
# clone the repo and ship scaler.pkl alongside don't need the raw data.
TRAIN_DATA_ROOT = Path(os.environ.get("GOALBALL_TRAIN_ROOT",
                       r"C:\Users\USER\Desktop\Thesis\Paralkympics2024"))
TRAIN_GOALS_DIR = Path(os.environ.get("GOALBALL_GOALS_DIR",
                       r"C:\Users\USER\Desktop\Thesis\Goals_Paralympics\outputs"))
TRAIN_GAMES = ["ISR_-_CAN_5-1","TUR_-_BRA_3-1","TUR_-_ISR_5-4",
               "ISR_-_BRA_8-4","CHI_-_TUR_7-5","BRA_-_TUR_3-3"]

# Default to CUDA when a GPU is available; fall back to CPU so the script
# also runs on machines without a GPU.
if torch.cuda.is_available():
    DEVICE = "cuda"
    log(f"[device] CUDA available -> using GPU: {torch.cuda.get_device_name(0)}")
else:
    DEVICE = "cpu"
    log("[device] CUDA not available -> falling back to CPU")

# Feature definitions - MUST match train_final.py exactly
BALL_COLS     = ["ball_x","ball_y","ball_w","ball_h","ball_conf"]
THROWER_COLS  = ["thrower_x","thrower_y","thrower_w","thrower_h","thrower_conf"]
DEFENDER_COLS = ["defender_x","defender_y","defender_w","defender_h","defender_conf"]
FEATURE_COLS  = ["segment_type","rel_t","gap","defender_seen","thrower_seen","ball_seen",
    "ball_x","ball_y","ball_dx","ball_dy","ball_w","ball_h","ball_conf",
    "thrower_x","thrower_y","thrower_w","thrower_h","thrower_conf",
    "defender_x","defender_y","defender_w","defender_h","defender_conf"]
CLS_TO_OUTCOME = {0:'o', 1:'o', 2:'g', 3:'g', 4:'b', 5:'b'}
# Full class labels: <outcome><validity>. Validity 1 = LSTM thinks segment is a real
# throw; 0 = LSTM thinks it's a false positive (mainly trained on 'o0' segments).
CLS_LABELS     = {0:'o1', 1:'o0', 2:'g1', 3:'g0', 4:'b1', 5:'b0'}
FALSE_LABELS   = {'o0', 'g0', 'b0'}  # what to drop when DROP_FALSE_THROWS = True
OUTCOME_NAMES  = {'g':'goal', 'b':'block', 'o':'out'}

# ===================================================================
#  USER INPUTS
# ===================================================================
game    = input("Game name (output folder, e.g. TUR_-_BRA_3-1): ").strip().strip('"').strip("'")
out_dir = SCRIPT_DIR / "Pipeline_Outputs" / game
out_dir.mkdir(parents=True, exist_ok=True)
excel_path   = out_dir / f"{game}_Throws_data_predicted.xlsx"
inf_csv_path = out_dir / f"{game}_Throws_lstm_inference.csv"

_video_raw = input(
    "Video path  (or 'LIVE' / camera index 0, 1, 2… for live camera): "
).strip().strip('"').strip("'")
IS_LIVE   = _video_raw.upper() == "LIVE" or _video_raw.isdigit()
cam_index = (0 if _video_raw.upper() == "LIVE" else int(_video_raw)) if IS_LIVE else None
video_path = str(out_dir / f"{game}_live_recording.mp4") if IS_LIVE else _video_raw
if IS_LIVE:
    log("[LIVE] Note: cv2 does not capture audio — Crowd Noise Score will be 0.0 for this run.")

# ===================================================================
#  LSTM ARCHITECTURE (must match train_final.py)
# ===================================================================
class ThrowLSTM(nn.Module):
    def __init__(self, n=23, h=128, l=2, c=6):
        super().__init__()
        self.lstm = nn.LSTM(n, h, l, batch_first=True, bidirectional=True, dropout=0.3)
        self.attn = nn.Sequential(nn.Linear(h*2,64), nn.Tanh(), nn.Dropout(0.2), nn.Linear(64,1))
        self.head = nn.Sequential(nn.Linear(h*2,128), nn.ReLU(), nn.Dropout(0.4), nn.Linear(128,c))

    def forward(self, x, lengths):
        lc = lengths.to(dtype=torch.int64, device="cpu")
        packed = pack_padded_sequence(x, lc, batch_first=True, enforce_sorted=False)
        out, _ = self.lstm(packed)
        out, _ = pad_packed_sequence(out, batch_first=True)
        B, L, _ = out.size()
        idxs = torch.arange(L, device=out.device).unsqueeze(0).expand(B, L)
        mask = idxs >= lc.to(out.device).unsqueeze(1)
        scores = self.attn(out).squeeze(-1).masked_fill(mask, float("-inf"))
        context = (out * torch.softmax(scores, dim=1).unsqueeze(-1)).sum(dim=1)
        return self.head(context)

# ===================================================================
#  PREPROCESSING (mirrors train_final.clean_throw, outcome optional)
# ===================================================================
def clean_throw(df, W, H):
    df = df.copy()
    df = df.sort_values("frame").reset_index(drop=True)
    df["ball_seen"]     = (~df["ball_conf"].isna()).astype(int)
    df["thrower_seen"]  = (~df["thrower_conf"].isna()).astype(int)
    df["defender_seen"] = (~df["defender_conf"].isna()).astype(int)
    for cols in (BALL_COLS, THROWER_COLS, DEFENDER_COLS):
        df[cols] = df[cols].ffill().bfill()
    for a in ("ball","thrower","defender"):
        df[f"{a}_x"] /= W; df[f"{a}_w"] /= W
        df[f"{a}_y"] /= H; df[f"{a}_h"] /= H
    df["rel_t"] = df.index/(len(df)-1) if len(df) > 1 else 0.0
    df["gap"]   = df["frame"].diff().fillna(0)
    df[["ball_dx","ball_dy"]] = df[["ball_x","ball_y"]].diff().fillna(0)
    df["segment_type"] = (df["segment_type"]=="to").astype(int)
    return df

# ===================================================================
#  SCALER (cached) - re-fit from training data on first run
# ===================================================================
def find_video(p):
    gd = p.parent.parent
    v = list(gd.glob(f"{gd.name}*.mp4")) + list(gd.glob(f"{gd.name}*.mov"))
    return v[0] if v else None

def get_dims(vp, fb=(1920,1080)):
    if vp is None: return fb
    cap = cv2.VideoCapture(str(vp))
    w, h = int(cap.get(3)), int(cap.get(4))
    cap.release()
    return (w, h) if w and h else fb

def fit_scaler():
    if SCALER_CACHE.exists():
        log(f"[scaler] loading cached -> {SCALER_CACHE.name}")
        with open(SCALER_CACHE, "rb") as f: return pickle.load(f)
    log("[scaler] cache miss - fitting from training data (one-off)...")
    parts = []
    csvs = [TRAIN_DATA_ROOT/g/"outputs"/f"{g}_Throws_lstm_training.csv" for g in TRAIN_GAMES]
    csvs += sorted(TRAIN_GOALS_DIR.glob("*.csv"))
    for csv in csvs:
        if not csv.exists():
            print(f"  WARN missing {csv}"); continue
        W, H = get_dims(find_video(csv))
        raw = pd.read_csv(csv)
        raw.loc[raw["label"]==0, "outcome"] = raw.loc[raw["label"]==0, "outcome"].fillna("o")
        tag = csv.stem.split("_Throws")[0]
        raw["throw_uid"] = ((raw["segment_id"]+1)//2).astype(str).radd(f"{tag}_")
        df = raw.groupby("throw_uid", group_keys=False).apply(clean_throw, W=W, H=H)
        parts.append(df)
    df_all = pd.concat(parts, ignore_index=True)
    df_all[FEATURE_COLS] = df_all[FEATURE_COLS].fillna(0)
    sc = StandardScaler()
    sc.fit(df_all[FEATURE_COLS])
    with open(SCALER_CACHE, "wb") as f: pickle.dump(sc, f)
    log(f"[scaler] saved -> {SCALER_CACHE.name}")
    return sc

# ===================================================================
#  YAMNET SETUP  (optional — graceful fallback if TF not installed)
# ===================================================================
try:
    import tensorflow as tf
    import tensorflow_hub as hub
    _YAMNET_AVAILABLE = True
except ImportError:
    _YAMNET_AVAILABLE = False
    log("[YAMNet] tensorflow / tensorflow-hub not installed -> Crowd Noise Score will be 0.0")
    log("[YAMNet] Install with:  pip install tensorflow tensorflow-hub")

def _find_ffmpeg():
    """Return ffmpeg executable path, searching PATH and common conda locations."""
    import shutil, sys
    # 1. Standard PATH lookup
    ff = shutil.which("ffmpeg")
    if ff: return ff
    # 2. Conda env bin (works when script is run with full python path, bypassing PATH)
    env_bin = Path(sys.executable).parent
    for candidate in [
        env_bin / "ffmpeg.exe",
        env_bin / "ffmpeg",
        env_bin.parent / "Library" / "bin" / "ffmpeg.exe",  # conda Windows layout
    ]:
        if candidate.exists(): return str(candidate)
    return None

def _extract_audio_16k(video_path, wav_path):
    """Use ffmpeg to extract mono 16 kHz WAV from video. Returns True on success."""
    ff = _find_ffmpeg()
    if ff is None:
        log("[YAMNet] ffmpeg not found. Install with:  conda install -c conda-forge ffmpeg")
        return False
    try:
        subprocess.run(
            [ff, "-y", "-i", str(video_path),
             "-vn", "-ac", "1", "-ar", "16000",
             "-acodec", "pcm_s16le", str(wav_path)],
            check=True, capture_output=True
        )
        return True
    except subprocess.CalledProcessError as e:
        log(f"[YAMNet] ffmpeg error: {e.stderr.decode(errors='ignore')[:200]}")
        return False

def _load_wav_mono16k(wav_path):
    """Read a mono 16-bit PCM WAV into a float32 numpy array."""
    with wave.open(str(wav_path), 'r') as wf:
        raw = wf.readframes(wf.getnframes())
    return np.frombuffer(raw, dtype=np.int16).astype(np.float32) / 32768.0

def _find_crowd_class_indices(yamnet_model):
    """Return list of YAMNet class indices matching YAMNET_CROWD_KEYWORDS."""
    class_map_path = yamnet_model.class_map_path().numpy().decode()
    class_df = pd.read_csv(class_map_path)
    names = class_df["display_name"].str.lower().tolist()
    indices = [i for i, n in enumerate(names)
               if any(kw in n for kw in YAMNET_CROWD_KEYWORDS)]
    log(f"[YAMNet] crowd classes ({len(indices)}): "
        + ", ".join(class_df['display_name'].iloc[indices].tolist()))
    return indices

# ===================================================================
#  LOAD MODELS + SCALER
# ===================================================================
log(f"[YOLO] loading {YOLO_WEIGHTS}")
yolo = YOLO(YOLO_WEIGHTS)
log(f"[LSTM] loading {LSTM_WEIGHTS}")
lstm = ThrowLSTM().to(DEVICE)
lstm.load_state_dict(torch.load(LSTM_WEIGHTS, map_location=DEVICE))
lstm.eval()
scaler = fit_scaler()

# ===================================================================
#  CALIBRATION FRAME + USER INPUTS (teams, halftime)
# ===================================================================
if IS_LIVE:
    _cap_live = cv2.VideoCapture(cam_index)
    if not _cap_live.isOpened():
        raise RuntimeError(f"Cannot open camera index {cam_index}")
    fps = _cap_live.get(cv2.CAP_PROP_FPS) or 30.0
    W   = int(_cap_live.get(cv2.CAP_PROP_FRAME_WIDTH))
    H   = int(_cap_live.get(cv2.CAP_PROP_FRAME_HEIGHT))
    ret, calib = _cap_live.read()
    if not ret: raise RuntimeError("Cannot read first frame from camera")
    log(f"[LIVE] Camera opened — FPS={fps:.2f}  Res={W}x{H}")
    halftime_frame = None   # set at runtime when T is pressed
else:
    _cap_live = None
    _cap_tmp = cv2.VideoCapture(video_path)
    if not _cap_tmp.isOpened(): raise RuntimeError(f"Cannot open {video_path}")
    ret, calib = _cap_tmp.read()
    fps = _cap_tmp.get(cv2.CAP_PROP_FPS)
    W   = int(_cap_tmp.get(cv2.CAP_PROP_FRAME_WIDTH))
    H   = int(_cap_tmp.get(cv2.CAP_PROP_FRAME_HEIGHT))
    _cap_tmp.release()
    if not ret: raise RuntimeError("Failed to read calibration frame")
    log(f"[INFO] FPS={fps:.2f}  Res={W}x{H}")

cv2.namedWindow("Calibration Frame", cv2.WINDOW_NORMAL)
cv2.imshow("Calibration Frame", calib)
print("Press any key once you have reviewed the calibration frame...")
cv2.waitKey(0); cv2.destroyWindow("Calibration Frame")

lower_team = input("Enter Lower Team name (e.g., ISR): ")
upper_team = input("Enter Upper Team name (e.g., CAN): ")

if IS_LIVE:
    log("[LIVE] Press T during detection to mark halftime (teams will switch sides).")
else:
    halftime_str = input("Enter halftime time (hh:mm:ss or mm:ss): ")
    parts = halftime_str.split(':')
    if len(parts) == 2:
        hours = 0
        minutes, seconds = map(int, parts)
    elif len(parts) == 3:
        hours, minutes, seconds = map(int, parts)
    else:
        raise ValueError("Invalid time format. Use hh:mm:ss or mm:ss")
    total_seconds = hours*3600 + minutes*60 + seconds
    halftime_frame = int(total_seconds * fps)
    log(f"[INFO] Halftime frame: {halftime_frame}")

# --- YAMNet: load model + extract full-video audio once ---
_yamnet_model      = None
_yamnet_crowd_idx  = []
_audio_waveform    = None
_audio_tmp_wav     = None

if _YAMNET_AVAILABLE:
    log("[YAMNet] loading model from TF Hub (first run downloads ~13 MB)...")
    _yamnet_model = hub.load("https://tfhub.dev/google/yamnet/1")
    _yamnet_crowd_idx = _find_crowd_class_indices(_yamnet_model)
    log(f"[YAMNet] crowd class indices: {_yamnet_crowd_idx}")
    _audio_tmp_wav = Path(tempfile.mktemp(suffix=".wav"))
    if IS_LIVE:
        log("[YAMNet] Live mode — microphone will be recorded during detection, then scored.")
    else:
        log(f"[YAMNet] extracting audio from video -> {_audio_tmp_wav} ...")
        if _extract_audio_16k(video_path, _audio_tmp_wav):
            _audio_waveform = _load_wav_mono16k(_audio_tmp_wav)
            log(f"[YAMNet] audio loaded: {len(_audio_waveform)/16000:.1f}s at 16 kHz  "
                f"| samples={len(_audio_waveform)}  min={_audio_waveform.min():.4f}  max={_audio_waveform.max():.4f}")
        else:
            log("[YAMNet] ffmpeg failed — ffmpeg must be on PATH. Crowd Noise Score will be 0.0")

_yamnet_debug_done = False   # log detailed info for the first throw only

def yamnet_crowd_score(t):
    """Score crowd noise over the throw's TO segment + YAMNET_POST_SECONDS."""
    global _yamnet_debug_done
    if _yamnet_model is None:
        return 0.0
    if _audio_waveform is None:
        return 0.0
    if t['to'] is None:
        return 0.0
    to_frames = t['to']['frames']
    start_sec  = to_frames[0]  / fps
    end_sec    = to_frames[-1] / fps + YAMNET_POST_SECONDS
    s = int(start_sec * 16000)
    e = int(min(end_sec * 16000, len(_audio_waveform)))
    clip = _audio_waveform[s:e]
    if not _yamnet_debug_done:
        log(f"[YAMNet][DEBUG] throw {t['num']}: audio window {start_sec:.2f}s–{end_sec:.2f}s  "
            f"clip_samples={len(clip)}  clip_max={clip.max():.4f}")
    if len(clip) < 3200:
        if not _yamnet_debug_done:
            log(f"[YAMNet][DEBUG] clip too short (<0.2s) — returning 0.0")
            _yamnet_debug_done = True
        return 0.0
    scores, _, _ = _yamnet_model(clip)
    mean_scores  = scores.numpy().mean(axis=0)
    if not _yamnet_crowd_idx:
        if not _yamnet_debug_done:
            log("[YAMNet][DEBUG] crowd index list is empty — returning 0.0")
            _yamnet_debug_done = True
        return 0.0
    crowd_score = float(max(mean_scores[i] for i in _yamnet_crowd_idx))
    if not _yamnet_debug_done:
        top5 = sorted(enumerate(mean_scores), key=lambda x: -x[1])[:5]
        log(f"[YAMNet][DEBUG] top-5 classes: {[(i, round(float(v),4)) for i,v in top5]}")
        log(f"[YAMNet][DEBUG] crowd score for throw {t['num']}: {crowd_score:.6f}")
        _yamnet_debug_done = True
    return crowd_score

# ---------- 8-click goal-zone calibration ----------
labels_calib = [
    "1) lower right bottom", "2) lower right top",
    "3) lower left bottom",  "4) lower left top",
    "5) upper left bottom",  "6) upper left top",
    "7) upper right bottom", "8) upper right top",
]
pts = []
def on_click(evt, x, y, flags, param):
    if evt == cv2.EVENT_LBUTTONDOWN and len(pts) < 8:
        pts.append((x, y))
        cv2.circle(calib, (x, y), 5, (0,0,255), -1)
        cv2.putText(calib, labels_calib[len(pts)-1], (x+5, y-5),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,0,255), 1)
        cv2.imshow("Calibrate", calib)

cv2.namedWindow("Calibrate", cv2.WINDOW_NORMAL)
cv2.imshow("Calibrate", calib)
cv2.setMouseCallback("Calibrate", on_click)
print("Click these 8 points in order:")
for l in labels_calib: print(" ", l)
while len(pts) < 8:
    if cv2.waitKey(1) & 0xFF == ord('q'): break
cv2.destroyAllWindows()

p1,p2,p3,p4,p5,p6,p7,p8 = [np.array(p) for p in pts]
low_rt, low_lt = p2, p4
up_lt,  up_rt  = p6, p8
lower_poly = np.array([p1,p2,p4,p3], dtype=np.int32)
upper_poly = np.array([p5,p6,p8,p7], dtype=np.int32)

def which_goal(pt):
    if cv2.pointPolygonTest(lower_poly, tuple(pt), False) >= 0: return 'lower'
    if cv2.pointPolygonTest(upper_poly, tuple(pt), False) >= 0: return 'upper'
    return 'lower' if pt[1] > calib.shape[0]/2 else 'upper'

def compute_zone(P, A, B):
    P, A, B = np.array(P), np.array(A), np.array(B)
    u = np.dot(P-A, B-A) / (np.dot(B-A, B-A) + 1e-6)
    u = float(np.clip(u, 0, 1))
    # u==1 would map to 10, and any stray negatives to 0 — clamp to [1, 9].
    return max(1, min(9, int(u*9) + 1))

def compute_mean_coord(coords, segment_type, n=10):
    if not coords: return np.array([np.nan, np.nan])
    sel = coords[:n] if segment_type=='from' else coords[-n:]
    return np.mean(sel, axis=0)

# ===================================================================
#  YOLO DETECTION + SEGMENT BUILDING
# ===================================================================
_PREVIEW_WIN = "YOLO Live  |  T=halftime  Q=stop" if IS_LIVE else "YOLO Live Preview"
_CLS_NAME  = {0: "thrower", 1: "defender", 32: "ball"}
_CLS_COLOR = {0: (0,255,0), 1: (0,0,255), 32: (255,255,0)}

if IS_LIVE:
    log("[LIVE] Detection started — T = halftime, Q = stop & analyse.")
    _rec_path = Path(video_path)
    _writer   = cv2.VideoWriter(str(_rec_path), cv2.VideoWriter_fourcc(*'mp4v'), fps, (W, H))
    cap = _cap_live
    live_preview = True
    cv2.namedWindow(_PREVIEW_WIN, cv2.WINDOW_NORMAL)
    # Parallel microphone recording for YAMNet
    _live_audio_chunks = []
    _live_audio_sr = 16000
    if _SOUNDDEVICE_AVAILABLE and _yamnet_model is not None:
        def _audio_cb(indata, frames, time_info, status):
            _live_audio_chunks.append(indata.copy())
        _audio_stream = sd.InputStream(samplerate=_live_audio_sr, channels=1,
                                       dtype='float32', callback=_audio_cb)
        _audio_stream.start()
        log("[LIVE] Microphone recording started.")
    else:
        _audio_stream = None
        if _yamnet_model is not None:
            log("[LIVE] sounddevice not installed — Crowd Noise Score will be 0.0  "
                "(run: pip install -r requirements.txt)")
else:
    log("[YOLO] running detection on full video...  (press H in the console to toggle live preview)")
    cap = cv2.VideoCapture(video_path)
    live_preview = False

seq, cur = [], {'label':None,'coords':[],'frames':[]}
gap, max_gap, fn = 0, 2, 0
frame_features = []

while True:
    ret, frame = cap.read()
    if not ret: break
    fn += 1

    if IS_LIVE:
        _writer.write(frame)
    elif not live_preview and _console_h_pressed():
        live_preview = True
        cv2.namedWindow(_PREVIEW_WIN, cv2.WINDOW_NORMAL)
        log("[preview] ON")

    res = yolo.predict(frame, save=False, conf=0.7, verbose=False)[0]
    cls   = res.boxes.cls.cpu().numpy()
    xy    = res.boxes.xywh.cpu().numpy()
    confs = res.boxes.conf.cpu().numpy()

    players = [c for c in cls if c in (0,1)]
    label = 1 if len(players) > 1 else (players[0] if players else None)

    feat = dict(frame=fn, label=label,
                ball_x=None,ball_y=None,ball_w=None,ball_h=None,ball_conf=None,
                thrower_x=None,thrower_y=None,thrower_w=None,thrower_h=None,thrower_conf=None,
                defender_x=None,defender_y=None,defender_w=None,defender_h=None,defender_conf=None)
    for c, b, cf in zip(cls, xy, confs):
        x, y, w, h = map(float, b)
        if   c == 32 and feat['ball_x']    is None: feat.update(ball_x=x,ball_y=y,ball_w=w,ball_h=h,ball_conf=float(cf))
        elif c == 0  and feat['thrower_x'] is None: feat.update(thrower_x=x,thrower_y=y,thrower_w=w,thrower_h=h,thrower_conf=float(cf))
        elif c == 1  and feat['defender_x']is None: feat.update(defender_x=x,defender_y=y,defender_w=w,defender_h=h,defender_conf=float(cf))
    frame_features.append(feat)

    balls = [b[:2] for c, b in zip(cls, xy) if c == 32]
    if label is not None and balls:
        gap = 0
        if cur['label'] is None: cur = {'label':label,'coords':[],'frames':[]}
        if cur['label'] == label:
            cur['coords'].append(balls[0]); cur['frames'].append(fn)
        else:
            seq.append({'type':'from' if cur['label']==0 else 'to',
                        'coords':cur['coords'],'frames':cur['frames']})
            cur = {'label':label,'coords':[balls[0]],'frames':[fn]}
    else:
        if cur['label'] is not None:
            gap += 1
            if gap > max_gap:
                seq.append({'type':'from' if cur['label']==0 else 'to',
                            'coords':cur['coords'],'frames':cur['frames']})
                cur = {'label':None,'coords':[],'frames':[]}; gap = 0

    if live_preview:
        disp = frame.copy()
        for c, b, cf in zip(cls, xy, confs):
            x, y, w, h = map(int, b)
            tl, br = (x-w//2, y-h//2), (x+w//2, y+h//2)
            color = _CLS_COLOR.get(int(c), (200,200,200))
            name  = _CLS_NAME.get(int(c), f"cls{int(c)}")
            cv2.rectangle(disp, tl, br, color, 2)
            cv2.putText(disp, f"{name} {float(cf):.2f}", (tl[0], tl[1]-5),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
        if IS_LIVE:
            ht_label = f"HT @ {halftime_frame}" if halftime_frame else "T=halftime"
            cv2.putText(disp, f"frame {fn}  |  {ht_label}  |  Q=stop",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,255), 2)
        else:
            cv2.putText(disp, f"frame {fn}  |  H to close preview",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,255), 2)
        cv2.imshow(_PREVIEW_WIN, disp)
        key = cv2.waitKey(1) & 0xFF
        if IS_LIVE:
            if key in (ord('t'), ord('T')) and halftime_frame is None:
                halftime_frame = fn
                log(f"[LIVE] Halftime marked at frame {fn}  ({fn/fps:.1f}s)")
            elif key in (ord('q'), ord('Q')):
                log(f"[LIVE] Q pressed — stopping at frame {fn}, starting analysis...")
                break
        elif key == ord('h'):
            live_preview = False
            cv2.destroyWindow(_PREVIEW_WIN)
            log("[preview] OFF")

if IS_LIVE:
    _writer.release()
    cap.release()
    cv2.destroyAllWindows()
    log(f"[LIVE] {fn} frames  ({fn/fps:.1f}s)  saved → {Path(video_path).name}")
    if halftime_frame is None:
        log("[LIVE] Warning: no halftime marked — treating entire recording as first half.")
        halftime_frame = fn + 1
    max_frames = fn
    # Save captured microphone audio and load for YAMNet
    if _audio_stream is not None:
        _audio_stream.stop()
        _audio_stream.close()
        if _live_audio_chunks:
            _live_pcm_f32 = np.concatenate(_live_audio_chunks, axis=0).flatten()
            with wave.open(str(_audio_tmp_wav), 'w') as _wf:
                _wf.setnchannels(1)
                _wf.setsampwidth(2)
                _wf.setframerate(_live_audio_sr)
                _wf.writeframes((_live_pcm_f32 * 32767).astype(np.int16).tobytes())
            _audio_waveform = _live_pcm_f32
            log(f"[YAMNet] Live audio captured: {len(_live_pcm_f32)/_live_audio_sr:.1f}s  "
                f"samples={len(_live_pcm_f32)}")
else:
    cap.release()
    if live_preview:
        cv2.destroyWindow(_PREVIEW_WIN)
    _cap_mf = cv2.VideoCapture(video_path)
    max_frames = int(_cap_mf.get(cv2.CAP_PROP_FRAME_COUNT)); _cap_mf.release()
if cur['label'] is not None and cur['frames']:
    seq.append({'type':'from' if cur['label']==0 else 'to',
                'coords':cur['coords'],'frames':cur['frames']})

# collapse + extend 'to' segments by 80 frames
collapsed, grp = [], None
for s in seq:
    if grp is None or grp['type'] != s['type']:
        grp = {'type':s['type'],'coords':s['coords'][:],'frames':s['frames'][:]}
        collapsed.append(grp)
    else:
        grp['coords'].extend(s['coords']); grp['frames'].extend(s['frames'])

for s in collapsed:
    if s['type'] == 'to':
        last = s['frames'][-1]
        s['frames'].extend(range(last+1, min(last+81, max_frames)))

segment_id_of = {id(s): i+1 for i, s in enumerate(collapsed)}

# ===================================================================
#  COLLECT PER-FRAME FEATURES (fully automatic - no user prompts)
# ===================================================================
log(f"[FEATURES] collecting per-frame features for {len(collapsed)} segments...")
training_records = []
for i, s in enumerate(collapsed):
    for f in s['frames']:
        ff = frame_features[f-1]
        rec = {'segment_id': i+1, 'segment_type': s['type'], 'frame': f, 'label': 1}
        rec.update({k:v for k,v in ff.items() if k not in ('frame','label')})
        training_records.append(rec)

pd.DataFrame(training_records).to_csv(inf_csv_path, index=False)
log(f"[CSV] Saved inference features -> {inf_csv_path}")

# ===================================================================
#  PAIR FROM -> TO into throws  (identical to yolo_cnn_predict_2.py)
# ===================================================================
throws = []
pending_from = None
for s in collapsed:
    if s['type'] == 'from':
        pending_from = s
    elif s['type'] == 'to':
        if pending_from is None:
            continue  # orphan 'to' before any 'from' - skip (YOLO artifact)
        throws.append({'num': len(throws)+1, 'from': pending_from,
                       'to': s, 'incomplete': False})
        pending_from = None
if pending_from is not None:
    throws.append({'num': len(throws)+1, 'from': pending_from,
                   'to': None, 'incomplete': True})

# ===================================================================
#  Merge "phantom split" duplicate throws (before running LSTM)
# ===================================================================
if MERGE_DUPLICATE_THROWS and len(throws) > 1:
    def _throwing_side(t):
        if t['from'] is None: return None
        thrower_xy = (np.nan, np.nan)
        for f in t['from']['frames'][:1]:
            ff = frame_features[f-1]
            if ff['thrower_x'] is not None:
                thrower_xy = (ff['thrower_x'], ff['thrower_y']); break
        if not np.isnan(thrower_xy[0]):
            return which_goal(thrower_xy)
        fc = compute_mean_coord(t['from']['coords'], 'from', n=5)
        return which_goal(fc) if not np.isnan(fc[0]) else None

    phantom_gap_frames = int(MERGE_PHANTOM_GAP_SECONDS * fps)
    dedup, dropped = [], []
    i = 0
    while i < len(throws):
        if i == len(throws) - 1:
            dedup.append(throws[i]); break
        a, b = throws[i], throws[i+1]
        is_dup = False
        if a['from'] and a['to'] and b['from'] and b['to']:
            # Phantom split: YOLO re-opened B's FROM mid-flight, so both FROM segments
            # start at nearly the same time. Measure FROM-start to FROM-start gap.
            # Using TO-end gap instead would give a false positive whenever the ball
            # stays visible across a missed intermediate throw (continuous tracking).
            from_gap = b['from']['frames'][0] - a['from']['frames'][0]
            if 0 <= from_gap <= phantom_gap_frames:
                same_team = _throwing_side(a) == _throwing_side(b) and _throwing_side(a) is not None
                if same_team:
                    is_dup = True
        if is_dup:
            dropped.append(a['num']); i += 1  # drop earlier (a), keep later (b) for next iter
        else:
            dedup.append(a); i += 1
    if dropped:
        log(f"[DEDUP] merged {len(dropped)} duplicate throw(s) (kept later of each pair, dropped earlier): {dropped}")
    else:
        log(f"[DEDUP] no phantom-split duplicates found ({len(throws)} throws retained)")
    throws = dedup

# ===================================================================
#  LSTM INFERENCE per throw
# ===================================================================
log(f"[LSTM] predicting outcomes for {len(throws)} throws...")

def predict_outcome(t):
    """Returns (outcome_name, softmax_confidence, class_label) or ('', 0.0, '') if incomplete."""
    if t['from'] is None or t['to'] is None:
        return '', 0.0, ''
    rows = []
    for seg, st in [(t['from'],'from'), (t['to'],'to')]:
        sid = segment_id_of[id(seg)]
        for f in seg['frames']:
            ff = frame_features[f-1]
            r = {'segment_id': sid, 'segment_type': st, 'frame': f}
            r.update({k:v for k,v in ff.items() if k not in ('frame','label')})
            rows.append(r)
    df = pd.DataFrame(rows)
    df = clean_throw(df, W, H)
    df[FEATURE_COLS] = df[FEATURE_COLS].fillna(0)
    feat = scaler.transform(df[FEATURE_COLS]).astype("float32")
    x = torch.tensor(feat).unsqueeze(0).to(DEVICE)
    L = torch.tensor([x.size(1)], dtype=torch.int64)
    with torch.no_grad():
        logits = lstm(x, L)
        probs = torch.softmax(logits, dim=1)
        conf, cls_id = probs.max(dim=1)
    cls_int = int(cls_id)
    return OUTCOME_NAMES[CLS_TO_OUTCOME[cls_int]], float(conf), CLS_LABELS[cls_int]

def segment_stats(t):
    """Returns (total_frames, ball_detection_rate) for the throw."""
    frames = []
    for seg in (t['from'], t['to']):
        if seg is not None: frames.extend(seg['frames'])
    if not frames:
        return 0, 0.0
    seen = sum(1 for f in frames if frame_features[f-1]['ball_x'] is not None)
    return len(frames), seen/len(frames)

throw_predictions = {t['num']: predict_outcome(t) for t in throws}
throw_stats       = {t['num']: segment_stats(t)   for t in throws}

if _yamnet_model is None:
    log("[YAMNet] WARNING: model not loaded — all Crowd Noise Scores will be 0.0")
    log("[YAMNet] Run:  pip install tensorflow tensorflow-hub  then re-run this script")
log(f"[YAMNet] scoring crowd noise for {len(throws)} throws...")
throw_crowd_scores = {t['num']: round(yamnet_crowd_score(t), 4) for t in throws}
if _yamnet_model is not None:
    scored = sum(1 for v in throw_crowd_scores.values() if v > 0)
    log(f"[YAMNet] done — {scored}/{len(throws)} throws have non-zero crowd score")

# Optionally drop throws the LSTM tagged as false-positive (o0/g0/b0)
if DROP_FALSE_THROWS:
    keep_nums = {n for n, (_, _, cl) in throw_predictions.items() if cl not in FALSE_LABELS}
    dropped = len(throws) - len(keep_nums)
    if dropped:
        log(f"[FILTER] dropping {dropped} throw(s) predicted as false-positive (o0/g0/b0); keeping {len(keep_nums)}")
    throws = [t for t in throws if t['num'] in keep_nums]

# ===================================================================
#  BUILD RESULTS  +  WRITE EXCEL
# ===================================================================
current_lower, current_upper = lower_team, upper_team
results = []

for t in throws:
    num = t['num']
    fc = compute_mean_coord(t['from']['coords'], 'from', n=5)

    # first-thrower fallback (same logic as yolo_cnn_predict_2.py)
    thrower_xy = (np.nan, np.nan)
    for f in t['from']['frames'][:1]:
        ff = frame_features[f-1]
        if ff['thrower_x'] is not None:
            thrower_xy = (ff['thrower_x'], ff['thrower_y']); break
    side = which_goal(thrower_xy) if not np.isnan(thrower_xy[0]) else which_goal(fc)

    end_fr = t['to']['frames'][-1] if t['to'] is not None else t['from']['frames'][-1]
    if end_fr > halftime_frame:
        current_lower, current_upper = upper_team, lower_team
    else:
        current_lower, current_upper = lower_team, upper_team
    throwing  = current_lower if side == 'lower' else current_upper
    defending = current_upper if side == 'lower' else current_lower

    if side == 'lower':
        fz = compute_zone(fc, low_rt, low_lt); fz_label = f"{fz} Lower"
    else:
        fz = compute_zone(fc, up_lt, up_rt);  fz_label = f"{fz} Upper"

    if t['to'] is None:
        tc = np.array([np.nan, np.nan], dtype=float); tz_label = np.nan
    else:
        tc = compute_mean_coord(t['to']['coords'], 'to', n=5)
        ds = 'upper' if side == 'lower' else 'lower'
        if ds == 'lower':
            tz = compute_zone(tc, low_rt, low_lt); tz_label = f"{tz} Lower"
        else:
            tz = compute_zone(tc, up_lt, up_rt);  tz_label = f"{tz} Upper"

    outcome, conf, cls_label = throw_predictions.get(num, ('', 0.0, ''))
    n_frames, ball_rate = throw_stats.get(num, (0, 0.0))
    start_seconds = t['from']['frames'][0] / fps
    start_time_str = f"{int(start_seconds // 60):02d}:{int(start_seconds % 60):02d}"
    results.append({
        'Throw Number':         num,
        'Start Time (mm:ss)':   start_time_str,
        'From Zone':            fz_label,
        'To Zone':              tz_label,
        'Throwing Team':        throwing,
        'Defending Team':       defending,
        'Predicted Outcome':    outcome,
        'Predicted Class':      cls_label,
        'Prediction Confidence': round(conf, 3),
        'Throw Length (frames)': n_frames,
        'Ball Detection Rate':  round(ball_rate, 3),
        'Crowd Noise Score':    throw_crowd_scores.get(num, 0.0),
        'From Coord':           fc.tolist(),
        'To Coord':             tc.tolist(),
    })

pd.DataFrame(results).to_excel(excel_path, index=False)

# -------------------------------------------------------------------
#  Make Start Time column clickable -> opens VLC at that second
# -------------------------------------------------------------------
from openpyxl import load_workbook
from openpyxl.styles import Font

links_dir = out_dir / "links"
links_dir.mkdir(exist_ok=True)
abs_video = str(Path(video_path).resolve())

def _launcher_bat(video, secs):
    # Tries vlc on PATH, then default install locations. Falls through with a helpful message.
    return (
        "@echo off\r\n"
        "where vlc >nul 2>nul\r\n"
        f"if %errorlevel% equ 0 (start \"\" vlc \"{video}\" --start-time={secs} --play-and-exit & exit /b)\r\n"
        f"if exist \"C:\\Program Files\\VideoLAN\\VLC\\vlc.exe\" (start \"\" \"C:\\Program Files\\VideoLAN\\VLC\\vlc.exe\" \"{video}\" --start-time={secs} --play-and-exit & exit /b)\r\n"
        f"if exist \"C:\\Program Files (x86)\\VideoLAN\\VLC\\vlc.exe\" (start \"\" \"C:\\Program Files (x86)\\VideoLAN\\VLC\\vlc.exe\" \"{video}\" --start-time={secs} --play-and-exit & exit /b)\r\n"
        "echo VLC not found. Install from https://www.videolan.org/vlc/\r\n"
        "pause\r\n"
    )

for r in results:
    mm, ss = r['Start Time (mm:ss)'].split(':')
    secs = int(mm)*60 + int(ss)
    bat = links_dir / f"throw_{int(r['Throw Number']):03d}.bat"
    bat.write_text(_launcher_bat(abs_video, secs), encoding="utf-8")

wb = load_workbook(excel_path)
ws = wb.active
headers = {c.value: c.column for c in ws[1]}
start_col = headers['Start Time (mm:ss)']
link_font = Font(color="0563C1", underline="single")
for i, r in enumerate(results, start=2):
    bat = links_dir / f"throw_{int(r['Throw Number']):03d}.bat"
    cell = ws.cell(row=i, column=start_col)
    cell.hyperlink = str(bat.resolve())
    cell.font = link_font
wb.save(excel_path)

log(f"[DONE] Excel saved -> {excel_path}  (Start Time cells link to VLC)")

# Clean up temp audio file
if _audio_tmp_wav is not None and _audio_tmp_wav.exists():
    _audio_tmp_wav.unlink()

# ===================================================================
#  OPTIONAL: throw-by-throw playback with bounding boxes
# ===================================================================
if SHOW_THROWS and throws:
    log("[VIEW] Playing each throw. Press any key for next, 'q' to quit.")
    play_cap = cv2.VideoCapture(video_path)
    cv2.namedWindow("Throw Playback", cv2.WINDOW_NORMAL)
    frame_delay_ms = max(1, int(1000 / fps))  # natural-speed playback

    for t in throws:
        outcome, conf, _ = throw_predictions.get(t['num'], ('', 0.0, ''))
        all_frames = []
        if t['from'] is not None: all_frames.extend(t['from']['frames'])
        if t['to']   is not None: all_frames.extend(t['to']['frames'])
        if not all_frames: continue

        quit_view = False
        for f in all_frames:
            play_cap.set(cv2.CAP_PROP_POS_FRAMES, f - 1)
            ret, img = play_cap.read()
            if not ret: continue
            ff = frame_features[f-1]

            if ff['ball_x'] is not None:
                cx, cy = int(ff['ball_x']), int(ff['ball_y'])
                cv2.circle(img, (cx, cy), 8, (255,255,0), -1)
                cv2.putText(img, "ball", (cx+10, cy+10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,0), 2)
            if ff['thrower_x'] is not None:
                x, y, w, h = ff['thrower_x'], ff['thrower_y'], ff['thrower_w'], ff['thrower_h']
                tl, br = (int(x-w/2), int(y-h/2)), (int(x+w/2), int(y+h/2))
                cv2.rectangle(img, tl, br, (0,255,0), 2)
                cv2.putText(img, "thrower", (tl[0], tl[1]-5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,0), 2)
            if ff['defender_x'] is not None:
                x, y, w, h = ff['defender_x'], ff['defender_y'], ff['defender_w'], ff['defender_h']
                tl, br = (int(x-w/2), int(y-h/2)), (int(x+w/2), int(y+h/2))
                cv2.rectangle(img, tl, br, (0,0,255), 2)
                cv2.putText(img, "defender", (tl[0], tl[1]-5),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,255), 2)

            header = f"Throw {t['num']}/{len(throws)}  |  Predicted: {outcome}  (conf={conf:.2f})"
            cv2.putText(img, header, (10, 30),
                        cv2.FONT_HERSHEY_DUPLEX, 0.9, (0,255,255), 2)
            cv2.imshow("Throw Playback", img)
            if cv2.waitKey(frame_delay_ms) & 0xFF == ord('q'):
                quit_view = True; break

        if quit_view: break
        print(f"  Throw {t['num']}: {outcome} (conf={conf:.2f}) - press any key for next, 'q' to quit")
        if cv2.waitKey(0) & 0xFF == ord('q'): break

    play_cap.release()
    cv2.destroyAllWindows()
    log("[VIEW] Done.")
